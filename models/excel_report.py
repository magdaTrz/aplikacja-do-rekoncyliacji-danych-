from zipfile import BadZipFile

import numpy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from models.base import ObservableModel
from text_variables import TextEnum
import win32com.client as win32
from pydispatch import dispatcher

import pandas
import time

UPDATE_TEXT_SIGNAL = 'update_text'


class ExcelReport(ObservableModel):
    def __init__(self, file_name, stage, password: str | None = None):
        super().__init__()
        self.workbook = None
        self.password_report: str | None = password
        self.file_name = file_name
        self.stage = stage
        self.load_workbook()
        self.merge_statistics_dataframe: pandas.DataFrame = None
        self.percent_reconciliation_dataframe: pandas.DataFrame = None
        self.summary_dataframe: pandas.DataFrame = None
        self.sample_dataframe: pandas.DataFrame = None

    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_name)
        except (FileNotFoundError, BadZipFile):
            self.create_new_workbook()
        if self.password_report is not None:
            self.workbook.password = self.password_report

    def create_new_workbook(self):
        try:
            self.workbook = openpyxl.Workbook()
        except Exception as e:
            print(f"ExcelReport(): Error tworzę nowy excel : {e}")
        else:
            self.sheet = self.workbook.active

    def save_workbook(self):
        try:
            self.workbook.save(self.file_name)
        except Exception as e:
            print(f"ExcelReport():Error podczas zapisywania Excela: {e}")

    def merge_and_compare(self, dataframe_1: pandas.DataFrame,
                          dataframe_2: pandas.DataFrame, on_cols: list, compare_cols: list, round_cols: None = None):
        dispatcher.send(signal=UPDATE_TEXT_SIGNAL, message=f"Tworzę ramkę danych dla raportu field to field", head='info')
        if self.stage == TextEnum.LOAD:
            suffixes_tuple = ("_src", "_ext")
            _merge_mapping_dict = {
                "left_only": "wiersz tylko w SRC",
                "right_only": "wiersz tylko w EXT",
                "both": " ",
            }
        else:
            suffixes_tuple = ("_ext", "_tgt")
            _merge_mapping_dict = {
                "left_only": "wiersz tylko w EXT",
                "right_only": "wiersz tylko w TGT",
                "both": " ",
            }
        merged_df = pandas.merge(
            dataframe_1,
            dataframe_2,
            on=on_cols,
            suffixes=suffixes_tuple,
            how="outer",
            indicator=True,
        )

        #  renaming _merge column
        merged_df["_merge"] = merged_df["_merge"].map(_merge_mapping_dict)
        cols_order = (
                on_cols
                + ["_merge"]
                + [col for col in merged_df.columns if col not in on_cols + ["_merge"]]
        )

        # comparing columns and determining True False
        for col in compare_cols:
            col_0 = f"{col}{suffixes_tuple[0]}"
            col_1 = f"{col}{suffixes_tuple[1]}"
            merged_df[col_0] = merged_df[col_0].fillna("")
            merged_df[col_1] = merged_df[col_1].fillna("")
            if merged_df[col_0].dtype == merged_df[col_1].dtype:
                print(
                    f" {col} tego samego typu: {merged_df[col_0].dtype=} i {merged_df[col_1].dtype=}"
                )
                merged_df[f"{col}_"] = numpy.where(
                    merged_df[col_0] == merged_df[col_1], True, False
                )
                # merged_df[f"{col}_"] = merged_df[col_0].equals(merged_df[col_1])
            else:
                print(
                    f" {col} różnego typu: {merged_df[col_0].dtype=} i {merged_df[col_1].dtype=}"
                )
                merged_df[f"{col}_"] = merged_df.apply(
                    lambda row: row[col_0] == row[col_1], axis=1
                )
        return merged_df

    def create_summary_dataframe(
            self, merged_result: pandas.DataFrame
    ) -> pandas.DataFrame:
        """A function that creates a dataframe with information about the number of true and false records and the
        reconciliation percentage "Kolumna": ,"Ilość True":, "Ilość False":, "Procent True":"""
        summary_dict = {
            "Kolumna": [],
            "Ilość False": [],
            "Ilość True": [],
            "Procent True": [],
        }

        for col in merged_result.columns:
            if col.endswith("_"):
                # Download the comparison column
                comparison_col = merged_result[col]

                # Count True and False
                true_count = comparison_col.sum()
                false_count = len(comparison_col) - true_count

                # Calculate the percentage of True and False over all rows
                total_rows = len(comparison_col)
                true_percent = (true_count / total_rows) * 100

                # Add information to the dictionary
                summary_dict["Kolumna"].append(col)
                summary_dict["Ilość True"].append(
                    self.format_number_with_space(true_count)
                )
                summary_dict["Ilość False"].append(
                    self.format_number_with_space(false_count)
                )
                summary_dict["Procent True"].append(
                    self.format_procentage(true_percent)
                )
        summary_dataframe = pandas.DataFrame(summary_dict)
        return summary_dataframe

    @staticmethod
    def create_row_count_summary(
            stage: str,
            dataframe1: pandas.DataFrame,
            dataframe2: pandas.DataFrame,
            merge_on=None,
            text_description=None or str,
    ) -> pandas.DataFrame:
        """A function that creates a dataframe with information about the number of records and description.
        "Liczba rekordów w xyz:" """
        if stage == TextEnum.LOAD:
            title1 = "Liczba rekordów w SRC: "
            title2 = "Liczba rekordów w EXT: "
        elif stage == TextEnum.END:
            title1 = "Liczba rekordów w EXT: "
            title2 = "Liczba rekordów w TGT: "
        else:
            title1 = "Liczba rekordów w : "
            title2 = "Liczba rekordów w : "

        summary_dataframe = pandas.DataFrame()
        if merge_on is not None:
            dataframe1 = dataframe1.drop_duplicates(subset=merge_on).reset_index(
                drop=True
            )
            dataframe2 = dataframe2.drop_duplicates(subset=merge_on).reset_index(
                drop=True
            )
        summary_dataframe[title1] = [dataframe1.shape[0]]
        summary_dataframe[title2] = [dataframe2.shape[0]]
        summary_dataframe["Opis raportu: "] = [text_description]
        return summary_dataframe

    def save_to_excel(
            self,
            dataframes_dict: dict[str, list[pandas.DataFrame]],
            merge_on: list,
            save_to_active_workbook=False,
    ):
        dispatcher.send(signal=UPDATE_TEXT_SIGNAL, message=f"Zapisuję dane do arkusza Excel",
                        head='info')

        def get_last_filled_row(worksheet):
            for row in worksheet.iter_rows(max_row=worksheet.max_row, min_row=1):
                for cell in row:
                    if cell.value is None:
                        return cell.row - 1
            return worksheet.max_row

        for sheet_name, dataframes_list in dataframes_dict.items():
            if "check_sum_umowy" in sheet_name:
                if "check_sum_umowy" not in self.workbook.sheetnames:
                    self.workbook.create_sheet(title="check_sum_umowy")
                worksheet = self.workbook["check_sum_umowy"]
            else:
                worksheet = self.workbook.create_sheet(title=sheet_name)
            start_row = get_last_filled_row(worksheet)
            start_row_list = []

            for dataframe in dataframes_list:
                if dataframe.empty:
                    print(f"Warning: DataFrame is empty for sheet {sheet_name}")
                    continue
                for row in dataframe_to_rows(dataframe, index=False, header=True):
                    worksheet.append(row)
                worksheet.append([])
                start_row_list.append(start_row)
                start_row += len(dataframe) + 2
            if "check_sum" in sheet_name:
                if sheet_name == "check_sum_umowy_status_umowy":
                    # special row list for check_sum_umowy to formate correct rows
                    self.format_dataframe_with_colors(
                        worksheet, [0, 4, 8, 12, 16, 21], dataframe
                    )
                else:
                    self.format_dataframe_with_colors(
                        worksheet, start_row_list, dataframe
                    )
            elif "f2f" in sheet_name:
                self.format_dataframe_with_colors(
                    worksheet, start_row_list, dataframe, merge_on
                )

        self.del_sheet("Sheet")
        is_saved = self.save_report()
        return is_saved

    def del_sheet(self, sheet_name: str) -> None:
        """Function that removes 'Sheet'"""
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.workbook.remove(sheet)
        return

    def save_report(self) -> bool:
        try:
            self.workbook.save(self.file_name)
            return True
        except PermissionError:
            return False

    def create_f2f_report(
            self,
            dataframe_1: pandas.DataFrame,
            dataframe_2: pandas.DataFrame,
            merge_on_cols: [list],
            compare_cols: [list],
            text_description: str | None = "",
    ):
        merge_and_compare_dataframe = self.merge_and_compare(
            dataframe_1, dataframe_2, merge_on_cols, compare_cols
        )
        summary_dataframe = self.create_summary_dataframe(merge_and_compare_dataframe)
        self.summary_dataframe = summary_dataframe

        row_count_dataframe = self.create_row_count_summary(
            self.stage,
            dataframe_1,
            dataframe_2,
            merge_on=merge_on_cols,
            text_description=text_description,
        )
        self.merge_statistics_dataframe = self.create_merge_statistics(merge_and_compare_dataframe)
        self.percent_reconciliation_dataframe = self.create_reconciliation_statistics(merge_and_compare_dataframe)
        self.sample_dataframe = self.create_sample_datafame(merge_and_compare_dataframe)

        return [row_count_dataframe, summary_dataframe, merge_and_compare_dataframe]

    def check_sum_count(
            self,
            dataframe_1: pandas.DataFrame,
            dataframe_2: pandas.DataFrame,
            column_to_counts: str,
            text_description="",
    ):
        if self.stage == "load":
            suffixes_tuple = ("_src", "_ext")
        elif self.stage == "end":
            suffixes_tuple = ("_ext", "_tgt")
        else:
            suffixes_tuple = ("_x", "_y")

        field_sum_1 = dataframe_1.shape[0]
        field_sum_2 = dataframe_2.shape[0]

        summ_dic = {
            "": column_to_counts,
            f"count{suffixes_tuple[0]}": self.format_number_with_space(field_sum_1),
            f"count{suffixes_tuple[1]}": self.format_number_with_space(field_sum_2),
            "Różnica": self.format_number_with_space(abs(field_sum_1 - field_sum_2)),
            "Decyzja": "",
            "Komentarz": "",
        }
        summ_df = pandas.DataFrame(summ_dic, index=[0])
        df = pandas.DataFrame()
        count_dic = {
            "": "",
            "Opis:": text_description,
        }
        count_df = pandas.DataFrame(count_dic, index=[0])
        return [count_df, summ_df, df]

    def check_sum(
            self,
            dataframe1: pandas.DataFrame,
            dataframe2: pandas.DataFrame,
            column_to_counts: str,
            text_description="",
    ):
        dispatcher.send(signal=UPDATE_TEXT_SIGNAL, message=f"Tworzę ramkę danych dla raportu check sum: {column_to_counts}",
                        head='info')

        if self.stage == TextEnum.LOAD:
            suffixes_tuple = ("_src", "_ext")
        elif self.stage == TextEnum.END:
            suffixes_tuple = ("_ext", "_tgt")
        else:
            suffixes_tuple = ("_x", "_y")
        counted_values_df1 = (
            dataframe1[column_to_counts].value_counts().to_frame().reset_index()
        )
        counted_values_df2 = (
            dataframe2[column_to_counts].value_counts().to_frame().reset_index()
        )
        merged_counted_values_df = pandas.merge(
            counted_values_df1,
            counted_values_df2,
            on=column_to_counts,
            suffixes=suffixes_tuple,
            how="outer",
        )
        merged_counted_values_dict = merged_counted_values_df.to_dict("records")
        count_1 = []
        count_2 = []
        absolute_diff = []
        type_name = []
        count_suffixes_str1 = f"count{suffixes_tuple[0]}"
        count_suffixes_str2 = f"count{suffixes_tuple[1]}"
        for item in merged_counted_values_dict:
            if numpy.isnan(item[count_suffixes_str1]):
                item[count_suffixes_str1] = 0
            if numpy.isnan(item[count_suffixes_str2]):
                item[count_suffixes_str2] = 0
            count_1.append(item[count_suffixes_str1])
            count_2.append(item[count_suffixes_str2])
            absolute_diff.append(
                abs(item[count_suffixes_str1] - item[count_suffixes_str2])
            )
            type_name.append(item[column_to_counts])
        summ_dic = {
            column_to_counts: type_name,
            count_suffixes_str1: count_1,
            count_suffixes_str2: count_2,
            "Różnica": self.format_diff_list(
                self.format_list_with_space(absolute_diff)
            ),
            "Decyzja": "",
            "Komentarz": "",
            f"Opis: {text_description}": ""
        }
        summ_df = pandas.DataFrame(summ_dic)
        return [summ_df]

    @staticmethod
    def add_password_to_excel(file_path: str,
                              password: str):
        try:
            dispatcher.send(signal=UPDATE_TEXT_SIGNAL, message=f"Nadaję hasło na raport rekoncyliacji",
                            head='info')
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workbook = excel.Workbooks.Open(file_path)
            workbook.Password = password
            workbook.Save()
            workbook.Close()
            excel.Quit()
            print(f"Password successfully added to {file_path}")
        except Exception as e:
            print(f"An error occurred: {e}")

    @staticmethod
    def format_procentage(percentage: float):
        return "{:.3f}%".format(percentage)

    @staticmethod
    def format_number_with_space(number: int):
        try:
            number_int = int(number)
            return "{:,}".format(number_int).replace(",", " ")
        except ValueError:
            if number == "nan":
                number = 0
            return number

    @staticmethod
    def format_list_with_space(number_list: list[int]):
        number_list = [0.0 if str(x).lower() == "nan" else x for x in number_list]
        return ["{:,.0f}".format(x).replace(",", " ") for x in number_list]

    @staticmethod
    def format_diff_list(list):
        return [("OK" if diff == "0" else diff) for diff in list]

    @staticmethod
    def format_dataframe_with_colors(
            worksheet,
            rows_to_format: list[int],
            dataframe: pandas.DataFrame,
            merge_on=None,
    ) -> None:
        last_row_to_format = max(rows_to_format) + 1

        fill = PatternFill(start_color="8ea9db", end_color="8ea9db", fill_type="solid")
        font_bold = Font(bold=True)
        font_gray = Font(color="808080")

        for row_idx in rows_to_format:
            for col_idx, col_name in enumerate(dataframe.columns):
                cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)
                cell.fill = fill
                cell.font = font_bold
        if merge_on is not None:
            columns_to_format = dataframe.columns[: len(merge_on)] if merge_on else []
            max_row = dataframe.shape[0] + last_row_to_format
            # format key as grey font
            for row in worksheet.iter_rows(
                    min_row=last_row_to_format,
                    max_row=max_row,
                    min_col=1,
                    max_col=len(columns_to_format),
            ):
                for cell in row:
                    cell.font = font_gray

    def create_merge_statistics(self, dataframe_merge: pandas.DataFrame) -> pandas.DataFrame:
        if self.stage == TextEnum.LOAD:
            suffixes_tuple = ("src", "ext")
            merge_tuple = ("wiersz tylko w SRC", "wiersz tylko w EXT")
        elif self.stage == TextEnum.END:
            suffixes_tuple = ("ext", "tgt")
            merge_tuple = ("wiersz tylko w EXT", "wiersz tylko w TGT")
        else:
            suffixes_tuple = ("błąd", "błąd")
            merge_tuple = ("błąd", "błąd")
        total_rows = len(dataframe_merge)
        only_1 = len(dataframe_merge[dataframe_merge["_merge"] == merge_tuple[0]])
        only_2 = len(dataframe_merge[dataframe_merge["_merge"] == merge_tuple[1]])
        both = len(dataframe_merge[dataframe_merge["_merge"] == " "])
        success_rate = (both / total_rows) * 100 if total_rows > 0 else 0

        stats = {
            "Index": [1, 2, 3, 4, 5, 6, 7],
            "Opis zmiennej": [
                "Wierszy",
                f"Tylko w {suffixes_tuple[0]}",
                f"Tylko w {suffixes_tuple[1]}",
                "Wiersze w obu",
                "Procent udanych połączeń (%)",
                f"Liczba wierszy w {suffixes_tuple[0]}",
                f"Liczba wierszy w {suffixes_tuple[1]}"
            ],
            "Wartość": [
                int(total_rows),
                int(only_1),
                int(only_2),
                both,
                success_rate,
                len(dataframe_merge[dataframe_merge['_merge'] != merge_tuple[1]]),
                len(dataframe_merge[dataframe_merge['_merge'] != merge_tuple[0]])
            ]
        }
        dataframe = pandas.DataFrame(stats)
        dataframe = dataframe.set_index('Index')
        return dataframe

    @staticmethod
    def create_reconciliation_statistics(dataframe_merge: pandas.DataFrame) -> pandas.DataFrame:
        total_rows = len(dataframe_merge)
        reconciliation_columns = [col for col in dataframe_merge.columns if col.endswith("_")]
        stats = {}

        for column in reconciliation_columns:
            true_count = dataframe_merge[column].sum()
            false_count = total_rows - true_count
            true_percentage = (true_count / total_rows) * 100 if total_rows > 0 else 0
            false_percentage = 100 - true_percentage

            stats[column] = {
                'Kolumna': column,
                'Liczba wierszy': int(total_rows),
                'True': int(true_count),
                'False': int(false_count),
                'True (%)': round(true_percentage, 2),
                'False (%)': round(false_percentage,2)
            }

        return pandas.DataFrame(stats).transpose()

    @staticmethod
    def create_sample_datafame(dataframe: pandas.DataFrame):
        if len(dataframe) > 100:
            return dataframe.head(100)
        else:
            return dataframe